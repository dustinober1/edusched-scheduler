"""Import/export functionality for EduSched."""

import json
import csv
from datetime import datetime, timedelta
from typing import List, Dict, Any, Union
from io import StringIO
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edusched.domain.problem import Problem
from edusched.domain.session_request import SessionRequest
from edusched.domain.resource import Resource
from edusched.domain.calendar import Calendar
from edusched.constraints.hard_constraints import *
from edusched.objectives.objectives import *


class DataFormatHandler:
    """Base class for data format handlers."""
    
    def import_data(self, data: Union[str, Dict, List]) -> Any:
        """Import data from the specific format."""
        raise NotImplementedError
    
    def export_data(self, data: Any) -> Union[str, bytes]:
        """Export data to the specific format."""
        raise NotImplementedError


class JSONHandler(DataFormatHandler):
    """Handler for JSON format."""
    
    def import_data(self, data: Union[str, Dict]) -> Problem:
        """Import problem from JSON string or dict."""
        if isinstance(data, str):
            data = json.loads(data)
        
        # Create the problem from the JSON data
        requests = []
        for req_data in data.get("requests", []):
            request = SessionRequest(
                id=req_data["id"],
                duration=timedelta(minutes=req_data["duration_minutes"]),
                number_of_occurrences=req_data["number_of_occurrences"],
                earliest_date=datetime.fromisoformat(req_data["earliest_date"]),
                latest_date=datetime.fromisoformat(req_data["latest_date"]),
                enrollment_count=req_data.get("enrollment_count", 0),
                min_capacity=req_data.get("min_capacity"),
                max_capacity=req_data.get("max_capacity"),
                required_attributes=req_data.get("required_attributes", {}),
                modality=req_data.get("modality", "in_person")
            )
            requests.append(request)
        
        resources = []
        for res_data in data.get("resources", []):
            resource = Resource(
                id=res_data["id"],
                resource_type=res_data["resource_type"],
                capacity=res_data.get("capacity"),
                attributes=res_data.get("attributes", {})
            )
            resources.append(resource)
        
        calendars = []
        for cal_data in data.get("calendars", []):
            calendar = Calendar(
                id=cal_data["id"],
                timezone=cal_data.get("timezone")
            )
            calendars.append(calendar)
        
        # For simplicity, we're not importing constraints and objectives here
        # In a full implementation, these would be handled as well
        
        return Problem(
            requests=requests,
            resources=resources,
            calendars=calendars,
            constraints=[],
            objectives=[]
        )
    
    def export_data(self, problem: Problem) -> str:
        """Export problem to JSON string."""
        data = {
            "requests": [
                {
                    "id": req.id,
                    "duration_minutes": int(req.duration.total_seconds() / 60),
                    "number_of_occurrences": req.number_of_occurrences,
                    "earliest_date": req.earliest_date.isoformat(),
                    "latest_date": req.latest_date.isoformat(),
                    "enrollment_count": req.enrollment_count,
                    "min_capacity": req.min_capacity,
                    "max_capacity": req.max_capacity,
                    "required_attributes": req.required_attributes,
                    "modality": req.modality
                }
                for req in problem.requests
            ],
            "resources": [
                {
                    "id": res.id,
                    "resource_type": res.resource_type,
                    "capacity": res.capacity,
                    "attributes": res.attributes
                }
                for res in problem.resources
            ],
            "calendars": [
                {
                    "id": cal.id,
                    "timezone": str(cal.timezone) if cal.timezone else None
                }
                for cal in problem.calendars
            ]
        }
        
        return json.dumps(data, indent=2)


class CSVHandler(DataFormatHandler):
    """Handler for CSV format."""
    
    def import_data(self, data: Union[str, StringIO]) -> Problem:
        """Import problem from CSV string or StringIO."""
        if isinstance(data, str):
            data = StringIO(data)
        
        # Read CSV data
        reader = csv.DictReader(data)
        
        requests = []
        resources = []
        calendars = []
        
        # Process different types of records based on a 'type' column
        for row in reader:
            record_type = row.get('type', '').lower()
            
            if record_type == 'request':
                request = SessionRequest(
                    id=row['id'],
                    duration=timedelta(minutes=int(row['duration_minutes'])),
                    number_of_occurrences=int(row['number_of_occurrences']),
                    earliest_date=datetime.fromisoformat(row['earliest_date']),
                    latest_date=datetime.fromisoformat(row['latest_date']),
                    enrollment_count=int(row.get('enrollment_count', 0)),
                    min_capacity=int(row.get('min_capacity', 0)) if row.get('min_capacity') else None,
                    max_capacity=int(row.get('max_capacity', 0)) if row.get('max_capacity') else None,
                    required_attributes=json.loads(row.get('required_attributes', '{}')),
                    modality=row.get('modality', 'in_person')
                )
                requests.append(request)
            
            elif record_type == 'resource':
                resource = Resource(
                    id=row['id'],
                    resource_type=row['resource_type'],
                    capacity=int(row.get('capacity', 0)) if row.get('capacity') else None,
                    attributes=json.loads(row.get('attributes', '{}'))
                )
                resources.append(resource)
            
            elif record_type == 'calendar':
                calendar = Calendar(
                    id=row['id'],
                    timezone=row.get('timezone')
                )
                calendars.append(calendar)
        
        return Problem(
            requests=requests,
            resources=resources,
            calendars=calendars,
            constraints=[],
            objectives=[]
        )
    
    def export_data(self, problem: Problem) -> str:
        """Export problem to CSV string."""
        output = StringIO()
        writer = None
        
        # Write requests
        for req in problem.requests:
            if writer is None:
                writer = csv.DictWriter(output, fieldnames=[
                    'type', 'id', 'duration_minutes', 'number_of_occurrences',
                    'earliest_date', 'latest_date', 'enrollment_count',
                    'min_capacity', 'max_capacity', 'required_attributes', 'modality'
                ])
                writer.writeheader()
            
            writer.writerow({
                'type': 'request',
                'id': req.id,
                'duration_minutes': int(req.duration.total_seconds() / 60),
                'number_of_occurrences': req.number_of_occurrences,
                'earliest_date': req.earliest_date.isoformat(),
                'latest_date': req.latest_date.isoformat(),
                'enrollment_count': req.enrollment_count,
                'min_capacity': req.min_capacity,
                'max_capacity': req.max_capacity,
                'required_attributes': json.dumps(req.required_attributes),
                'modality': req.modality
            })
        
        # Write resources
        for res in problem.resources:
            if writer is None:
                writer = csv.DictWriter(output, fieldnames=[
                    'type', 'id', 'resource_type', 'capacity', 'attributes'
                ])
                writer.writeheader()
            
            writer.writerow({
                'type': 'resource',
                'id': res.id,
                'resource_type': res.resource_type,
                'capacity': res.capacity,
                'attributes': json.dumps(res.attributes)
            })
        
        # Write calendars
        for cal in problem.calendars:
            if writer is None:
                writer = csv.DictWriter(output, fieldnames=[
                    'type', 'id', 'timezone'
                ])
                writer.writeheader()
            
            writer.writerow({
                'type': 'calendar',
                'id': cal.id,
                'timezone': str(cal.timezone) if cal.timezone else None
            })
        
        return output.getvalue()


class ExcelHandler(DataFormatHandler):
    """Handler for Excel format."""
    
    def import_data(self, file_path: str) -> Problem:
        """Import problem from Excel file."""
        workbook = openpyxl.load_workbook(file_path)
        
        requests = []
        resources = []
        calendars = []
        
        # Process requests sheet
        if 'Requests' in workbook.sheetnames:
            requests_sheet = workbook['Requests']
            headers = [cell.value for cell in requests_sheet[1]]
            
            for row in requests_sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                
                if row_dict.get('id'):
                    request = SessionRequest(
                        id=str(row_dict['id']),
                        duration=timedelta(minutes=int(row_dict['duration_minutes'] or 0)),
                        number_of_occurrences=int(row_dict['number_of_occurrences'] or 0),
                        earliest_date=datetime.fromisoformat(str(row_dict['earliest_date'])),
                        latest_date=datetime.fromisoformat(str(row_dict['latest_date'])),
                        enrollment_count=int(row_dict.get('enrollment_count', 0)),
                        min_capacity=int(row_dict.get('min_capacity')) if row_dict.get('min_capacity') else None,
                        max_capacity=int(row_dict.get('max_capacity')) if row_dict.get('max_capacity') else None,
                        required_attributes=json.loads(row_dict.get('required_attributes', '{}')) if row_dict.get('required_attributes') else {},
                        modality=row_dict.get('modality', 'in_person')
                    )
                    requests.append(request)
        
        # Process resources sheet
        if 'Resources' in workbook.sheetnames:
            resources_sheet = workbook['Resources']
            headers = [cell.value for cell in resources_sheet[1]]
            
            for row in resources_sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                
                if row_dict.get('id'):
                    resource = Resource(
                        id=str(row_dict['id']),
                        resource_type=str(row_dict['resource_type']),
                        capacity=int(row_dict.get('capacity')) if row_dict.get('capacity') else None,
                        attributes=json.loads(row_dict.get('attributes', '{}')) if row_dict.get('attributes') else {}
                    )
                    resources.append(resource)
        
        # Process calendars sheet
        if 'Calendars' in workbook.sheetnames:
            calendars_sheet = workbook['Calendars']
            headers = [cell.value for cell in calendars_sheet[1]]
            
            for row in calendars_sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                
                if row_dict.get('id'):
                    calendar = Calendar(
                        id=str(row_dict['id']),
                        timezone=row_dict.get('timezone')
                    )
                    calendars.append(calendar)
        
        return Problem(
            requests=requests,
            resources=resources,
            calendars=calendars,
            constraints=[],
            objectives=[]
        )
    
    def export_data(self, problem: Problem, file_path: str = None) -> Union[bytes, None]:
        """Export problem to Excel file."""
        workbook = Workbook()
        
        # Create requests sheet
        requests_sheet = workbook.active
        requests_sheet.title = "Requests"
        requests_sheet.append([
            'id', 'duration_minutes', 'number_of_occurrences', 'earliest_date', 
            'latest_date', 'enrollment_count', 'min_capacity', 'max_capacity', 
            'required_attributes', 'modality'
        ])
        
        for req in problem.requests:
            requests_sheet.append([
                req.id,
                int(req.duration.total_seconds() / 60),
                req.number_of_occurrences,
                req.earliest_date.isoformat(),
                req.latest_date.isoformat(),
                req.enrollment_count,
                req.min_capacity,
                req.max_capacity,
                json.dumps(req.required_attributes),
                req.modality
            ])
        
        # Create resources sheet
        resources_sheet = workbook.create_sheet("Resources")
        resources_sheet.append([
            'id', 'resource_type', 'capacity', 'attributes'
        ])
        
        for res in problem.resources:
            resources_sheet.append([
                res.id,
                res.resource_type,
                res.capacity,
                json.dumps(res.attributes)
            ])
        
        # Create calendars sheet
        calendars_sheet = workbook.create_sheet("Calendars")
        calendars_sheet.append([
            'id', 'timezone'
        ])
        
        for cal in problem.calendars:
            calendars_sheet.append([
                cal.id,
                str(cal.timezone) if cal.timezone else None
            ])
        
        if file_path:
            workbook.save(file_path)
            return None
        else:
            # Return as bytes
            from io import BytesIO
            stream = BytesIO()
            workbook.save(stream)
            stream.seek(0)
            return stream.getvalue()


class ImportExportManager:
    """Manages import/export operations."""
    
    def __init__(self):
        self.handlers = {
            'json': JSONHandler(),
            'csv': CSVHandler(),
            'excel': ExcelHandler(),
        }
    
    def import_from_format(self, data: Union[str, bytes], format_type: str) -> Problem:
        """Import data from the specified format."""
        if format_type not in self.handlers:
            raise ValueError(f"Unsupported format: {format_type}")
        
        if format_type == 'excel' and isinstance(data, str):
            # For Excel, data is typically a file path
            return self.handlers[format_type].import_data(data)
        else:
            return self.handlers[format_type].import_data(data)
    
    def export_to_format(self, problem: Problem, format_type: str, destination: str = None) -> Union[str, bytes, None]:
        """Export problem to the specified format."""
        if format_type not in self.handlers:
            raise ValueError(f"Unsupported format: {format_type}")
        
        result = self.handlers[format_type].export_data(problem)
        
        if destination:
            if format_type == 'excel':
                # For Excel, save to file
                self.handlers[format_type].export_data(problem, destination)
                return None
            else:
                # For text formats, write to file
                with open(destination, 'w') as f:
                    f.write(result)
                return None
        
        return result
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported formats."""
        return list(self.handlers.keys())